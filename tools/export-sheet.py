#!/usr/bin/env python3
"""
Regenerate dacss-website/data/list.json from dacss.xlsx.

Run this whenever Albert edits the DACSS LIST tab, then commit the new list.json.

    python3 tools/export-sheet.py ../dacss.xlsx

Requires: openpyxl  (pip install openpyxl)

How it reads the sheet:
  - Column A holds a category-index label (1.1, 1.2, ...) on disorder rows,
    and a category *name* on the header row that precedes each block.
  - Column B holds the disorder name; a trailing '^' means "reported after CO
    poisoning" and '*' means "central sensitization syndrome" (kept as flags).
  - Column C holds the status code (default N) — not exported; the website
    starts every item at N.

Stable IDs: ids are dacss-<category>-<NN>. If you REORDER or RENAME disorders,
re-run this and diff list.json so you can see adds/removes/renames before
committing — that keeps patients' saved answers matched to the right item.
"""
import sys, json, re, os

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# (id, display name, first data row, last data row) for each category block.
# Update these ranges if the sheet's structure changes.
CATEGORIES = [
    ('affective',        'Affective / Behavior / Mood / Psychiatric / Psychological', 18, 32),
    ('autoimmune',       'Autoimmune, Infectious, or Post-Infectious',                35, 58),
    ('endocrine',        'Endocrine / Metabolic',                                     61, 76),
    ('gastrointestinal', 'Gastrointestinal',                                          79, 88),
    ('musculoskeletal',  'Musculoskeletal and Pelvic',                                91, 102),
    ('neurological',     'Neurological',                                              105, 127),
    ('respiratory',      'Respiratory or ENT',                                        130, 142),
    ('toxic',            'Toxic Exposure (External or Internal)',                     145, 161),
]
VERSION = 'v26.08.23'  # bump to match the sheet's in-cell version


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else 'dacss.xlsx'
    if not os.path.exists(src):
        sys.exit(f"Workbook not found: {src}")

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb['DACSS LIST']

    out = {'report': 'DACSS', 'version': VERSION, 'categories': []}
    total = 0
    for cid, cname, r0, r1 in CATEGORIES:
        items = []
        for r in range(r0, r1 + 1):
            name = ws.cell(r, 2).value
            if name is None or not str(name).strip():
                continue
            name = str(name).strip()
            items.append({
                'id': f'dacss-{cid}-{len(items) + 1:02d}',
                'sourceRow': r,
                'displayIndex': str(ws.cell(r, 1).value),
                'name': name.replace('*', '').replace('^', '').strip(),
                'centralSensitization': '*' in name,
                'coPoisoning': '^' in name,
            })
        total += len(items)
        out['categories'].append({'id': cid, 'name': cname, 'items': items})
        print(f'  {cname}: {len(items)}')

    dest = os.path.join(os.path.dirname(__file__), '..', 'data', 'list.json')
    dest = os.path.normpath(dest)
    with open(dest, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f'\nWrote {dest}  ({total} disorders)')


if __name__ == '__main__':
    main()
